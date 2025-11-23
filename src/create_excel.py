import os
import pandas as pd
from datetime import datetime 
from persiantools import digits
from persiantools.jdatetime import JalaliDateTime
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

class ExcelTable:
	"""
    A class to generate a styled Excel file containing activity plans and mental/physical state data,
    with support for Persian and English languages.

    Attributes:
        data (list): List of activity entries (each entry is a list of estimated time and activity name).
        language (str): Language code ("Fa" for Persian, "En" for English).
        excel_path (str): Path to the output Excel file.
    """
	def __init__(self, data, lang="Fa", file_path=None):
		"""
        Initializes the ExcelTable object, builds internal DataFrames based on input data and language.

        Args:
            data (list): List of [estimated_time, activity_name] entries.
            lang (str): Language code ("Fa" or "En").
            file_path (str or None): Optional directory to save the Excel file. If None, uses current directory.
        """
		
		self.data = data
		self.language = lang
		
		if self.language=="En":
			self.df_states = pd.DataFrame({"Physical state":[10], "Mental state":[10]})
		# ADD Other languages like 2 following line
		# elif self.language=="En":
		# 	self.df_states = pd.DataFrame({"Physical state":[10], "Mental state":[10]})
		else:
			self.df_states = pd.DataFrame({"حالت جسمی":[10], "حالت روحی":[10]})



		self.df_data = pd.DataFrame()
		real_times = [""]* len(self.data)
		self.all_data = []
		for r, t in zip(real_times,self.data):
			self.all_data.append([r,t[1],t[0]])

		if self.language=="En":
			self.df_data =pd.DataFrame(self.all_data, columns=["Real Time","Time Estimate", "Activity Name"])
			self.df_data.loc[len(self.df_data)] = ["", "" ,"Sleep"]
		# ADD Other languages like 3 following line
		# elif self.language=="En":
		# 	self.df_data =pd.DataFrame(self.all_data, columns=["Real Time","Time Estimate", "Activity Name"])
		# 	self.df_data.loc[len(self.df_data)] = ["", "" ,"Sleep"]

		else:
			self.df_data =pd.DataFrame(self.all_data, columns=["زمان واقعی","تخمین حدودی", "لیست انجام کارها"])
			self.df_data.loc[len(self.df_data)] = ["", "" ,"خواب"]


		if file_path is None:
			file_path = os.path.dirname(os.path.realpath(__file__))

		if self.language == "En":
			date_str = datetime.now().strftime('%Y-%m-%d')
		else:
			date_str = JalaliDateTime.now().strftime('%Y-%m-%d') 

		self.excel_path = os.path.join(file_path, f"plan_{date_str}_{self.language}.xlsx")

		self.font_name = "Vazirmatn"
		self.palette_color = {'sleep': "5A9BD5", 'date': "F50206", 'header': "003494", 'border': "1C4E7C"}

	def create_excel(self):
		"""
        Creates an Excel file with styled tables including:
        - Activity schedule with estimated times
        - Physical and mental state entries
        - Date, header colors, borders, and fonts
        """
		with pd.ExcelWriter(self.excel_path, engine="openpyxl") as writer:

			self.df_data.to_excel(writer, index=False, sheet_name="Sheet1", startrow=1, startcol=3)
			self.df_states.to_excel(writer, index=False, sheet_name="Sheet1", startrow=1, startcol=1)

			wb = writer.book
			ws = wb.active

			# Set column widths
			ws.column_dimensions["B"].width = 15  # Mental State
			ws.column_dimensions["C"].width = 15  # Estimated Time
			ws.column_dimensions["D"].width = 15  # Actual Time
			ws.column_dimensions["E"].width = 15  # Task List
			ws.column_dimensions["F"].width = 25  # Task List
			ws.column_dimensions["G"].width = 15  # DateTime

			row_start = 2
			table1 = Table(displayName="Table1", ref=f"B{row_start}:C{self.df_states.shape[1]+1}")
			table2 = Table(displayName="Table2", ref=f"D{row_start}:F{self.df_data.shape[0]+1}")


			# Center-align all cells and apply font
			for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
				for cell in row:
					cell.alignment = Alignment(horizontal="center", vertical="center")
					cell.font = Font(name=self.font_name,size=11, color="000000")


			# Set date cell (Persian date, red background)
			date_cell = ws[f"G{row_start}"]
			if self.language == "En":
				date_cell.value = datetime.now().strftime('%Y.%m.%d') 
			else:
				date_cell.value = digits.en_to_fa(JalaliDateTime.now().strftime('%Y.%m.%d'))
			# date_cell.value = digits.en_to_fa(JalaliDateTime.now().strftime('%Y.%m.%d'))
			date_cell.fill = PatternFill(start_color=self.palette_color['date'], end_color=self.palette_color['date'], fill_type="solid")
			date_cell.border = Border(top=Side(style='medium', color=self.palette_color["border"]),
											left=Side(style='medium', color=self.palette_color["border"]),
											right=Side(style='medium', color=self.palette_color["border"]),
											bottom=Side(style='medium', color=self.palette_color["border"]))
			date_cell.font = Font(name=self.font_name,size=11 ,color="FFFFFF", bold=True)
			date_cell.alignment = Alignment(horizontal="center", vertical="center")

			# Name cell formatting (black background)
			name_cell = ws[f"F{row_start}"]
			name_cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
			name_cell.font = Font(name=self.font_name,color="FFFFFF", bold=True)
			name_cell.alignment = Alignment(horizontal="center", vertical="center")

			# Header styling
			header_fill = PatternFill(start_color=self.palette_color["header"], end_color=self.palette_color["header"], fill_type="solid")
			header_font = Font(name=self.font_name,color="FFFFFF", bold=True)
			
			for cell in ws[f"B{row_start}:E{row_start}"][0]:
				cell.fill = header_fill
				cell.font = header_font

			# Table style
			style = TableStyleInfo(
				name="TableStyleMedium2",
				showFirstColumn=False,
				showLastColumn=False,
				showRowStripes=True,
				showColumnStripes=False,
			)
			table1.tableStyleInfo = style
			table2.tableStyleInfo = style

			ws.add_table(table1)
			ws.add_table(table2)
			wb.save(self.excel_path)

			# Draw borders for both tables
			min_col = 2
			max_row = self.df_states.shape[0] + row_start  
			max_col = len(self.df_states.columns) + min_col -1
			self.border_table(ws,row_start, min_col, max_row, max_col, False)
			
			min_col = 4
			max_row = self.df_data.shape[0] + row_start
			max_col = len(self.df_data.columns)+min_col-1
			self.border_table(ws,row_start, min_col, max_row, max_col, True)

			wb.save(self.excel_path)

			# Remove auto filters for compatibility
			table1.autoFilter = None    
			table2.autoFilter = None    

			wb.save(self.excel_path)
			print(f"Excel file '{self.excel_path}' created successfully with a built-in table style!")

	def border_table (self, ws, min_row,min_col, max_row, max_col, has_sleep):
		"""
        Applies borders (and optionally sleep row coloring) to a table area in the worksheet.

        Args:
            ws (Worksheet): The openpyxl worksheet object.
            min_row (int): Starting row of the table.
            min_col (int): Starting column of the table.
            max_row (int): Ending row of the table.
            max_col (int): Ending column of the table.
            has_sleep (bool): Whether the last row is a sleep row and needs coloring.
        """
		for row in range(min_row, max_row + 1):
			for col in range(min_col, max_col + 1):
				cell = ws.cell(row=row, column=col)
				# Top border for header
				if row == min_row: 
					cell.border = Border(top=Side(style='medium', color=self.palette_color["border"]),
										left=Side(style='medium', color=self.palette_color["border"]),
										right=Side(style='medium', color=self.palette_color["border"]),
										bottom=Side(style='medium', color=self.palette_color["border"]))
				# Bottom border and fill for last row
				elif row == max_row:
					if col == min_col:
						cell.border = Border(bottom=Side(style='medium', color=self.palette_color["border"]),
											left=Side(style='medium', color=self.palette_color["border"]))
					elif col == max_col:
						cell.border = Border(bottom=Side(style='medium', color=self.palette_color["border"]),
											right=Side(style='medium', color=self.palette_color["border"]))
					else:
						cell.border = Border(bottom=Side(style='medium', color=self.palette_color["border"]))
					if has_sleep:
						cell.fill = PatternFill(start_color=self.palette_color['sleep'], end_color=self.palette_color['sleep'], fill_type="solid")
						cell.font = Font(name=self.font_name, size=11, color="FFFFFF")
				# Left border
				elif col == min_col:
					cell.border = Border(left=Side(style='medium', color=self.palette_color["border"]))
				# Right border
				elif col == max_col:
					cell.border = Border(right=Side(style='medium', color=self.palette_color["border"]))
